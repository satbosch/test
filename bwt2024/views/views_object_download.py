# -*- encoding: utf-8 -*-
"""
Copyright (c) 2024 - XC/EVI
"""

from django import template
from django.contrib import messages
from django.forms                       import *
from django.http import HttpResponse
from datetime import datetime   # datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.drawing.image import Image
from bwt2024.models                     import *
from bwt2024.forms                     import *
from mysite.settings import *
from openpyxl.drawing.image import Image
from django.template import loader
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Side
from datetime import datetime


# -- ==============================================================
# fonts
head_font               = Font(name='Arial', size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
head_white_font         = Font(name='Arial', size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FFFFFFFF')
data_font               = Font(name='Arial', size=10, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
data_font_bold          = Font(name='Arial', size=10, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
data_font_under         = Font(name='Arial', size=10, bold=False, italic=False, vertAlign=None, underline='single', strike=False, color='FF0000FF')

head_calibri            = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')
head_calibri_white      = Font(name='Calibri', size=11, bold=True, italic=False, vertAlign=None, underline='none', strike=False, color='FFFFFFFF')
head_verdana            = Font(name='Verdana', size=30, bold=False, italic=False, vertAlign=None, underline='none', strike=False, color='FF000000')


# -- ==============================================================
# alignment
head_align_center       = Alignment(vertical='center',horizontal='center',wrap_text=True)
data_align_center       = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
data_align_left         = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
data_align_left_wrap    = Alignment(horizontal='left', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
data_align_middle       = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
data_align_center_top   = Alignment(horizontal='center', vertical='top', text_rotation=0, wrap_text=True, shrink_to_fit=False, indent=0)
data_align_middle_rotate       = Alignment(horizontal='center', vertical='center', text_rotation=90, wrap_text=True, shrink_to_fit=False, indent=0)

# -- ==============================================================
# format
data_general_format     = 'General'
data_number_format      = 'Number'

# -- ==============================================================
# border
border_thin             = Side(border_style="thin", color="000000")
border_thin_white       = Side(border_style="thin", color="FFFFFF")
border_dashed           = Side(border_style="mediumDashed", color="000000")
border_square_thin      = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
border_square_dashed    = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_dashed)

border_square_thin_white = Border(top=border_thin_white, left=border_thin_white, right=border_thin_white, bottom=border_thin_white)

# -- ==============================================================
# color
vuln_scan_color_yellow  = PatternFill("solid", fgColor="FFFFCC")
vuln_scan_color_green   = PatternFill("solid", fgColor="CCFFCC")
vuln_scan_color_pink    = PatternFill("solid", fgColor="FFCCCC")
vuln_scan_color_salmon  = PatternFill("solid", fgColor="FFCF9E")
vuln_scan_color_l_blue  = PatternFill("solid", fgColor="CCFFFF")
vuln_scan_color_m_blue  = PatternFill("solid", fgColor="99CCFF")
vuln_scan_color_d_blue  = PatternFill("solid", fgColor="0066CC")
vuln_scan_color_o_blue  = PatternFill("solid", fgColor="33CCCC")

vuln_scan_color_pink         = PatternFill("solid", fgColor="A80163")
vuln_scan_color_yellow       = PatternFill("solid", fgColor="FFFF00")
vuln_scan_color_green        = PatternFill("solid", fgColor="00B050")
vuln_scan_color_red          = PatternFill("solid", fgColor="FF0000")
vuln_scan_color_orange       = PatternFill("solid", fgColor="FFC000")
vuln_scan_color_grey         = PatternFill("solid", fgColor="A5A5A5")

vuln_scan_color_odd          = PatternFill("solid", fgColor="E1CBD3")
vuln_scan_color_even         = PatternFill("solid", fgColor="F1E7EA")

vuln_scan_font               = Font(name="Bosch Office Sans", size=8, color="000000", bold=False)
vuln_scan_font_white         = Font(name="Bosch Office Sans", size=8, color="FFFFFF", bold=True)
vuln_scan_font_wingdings      = Font(name="Wingdings", size=8, color="000000", bold=False)



# -- ==============================================================
# height
data_height = 20


###########################################################
# Object_Unit Download
###########################################################
def object_unit_download(request):
    context = {}

    try:

        Object_Unit_FormSet = modelformset_factory(Object_Unit_Download, form=Object_Unit_Form, can_delete=True, fields = ('component','version') )
        context['formset'] = Object_Unit_FormSet(queryset=Object_Unit_Download.objects.all())

        # Save Button Logic

        if request.method=='POST' and 'btn_list_edit_save' in request.POST:
            formset = Object_Unit_FormSet(request.POST)
            if formset.is_valid():
                instances = formset.save(commit=False)
                for instance in instances:
                    try:
                        instance.save()
                        messages.success(request, 'Object Unit ' + instance.component + ' saved sucessfully.' )
                    except:
                        messages.error(request, 'Object Unit ' + instance.component + ' not saved.' )
                for obj in formset.deleted_objects:
                    try:
                        messages.success(request, 'Object Unit ' + obj.component + ' deleted sucessfully.' )
                        obj.delete()
                    except:
                        messages.error(request, 'Object Unit ' + obj.component + ' not deleted.' )
            else:
                messages.error(request, 'Error: Object Unit Form was not saved.' )
                for error in formset.errors:
                    messages.error(request, str(error) )

        # Download Excel Button Logic

        if request.method=='POST' and 'btn_download_excel' in request.POST:
            download = object_download(request.POST.get('btn_download_excel'))
            return download

        load_template           = "object_unit\object_unit_download.html"
        context['segment']      = load_template
        context['PageTitle']    = "Object Units Download"
        context['msghelp']      = "This is an Help message"
        context['breadcrumbs']  = [{'Name' : 'Home', 'Link' : '/', 'Active': False} , {'Name' : 'Object Units', 'Link' : '/Object_unit_form', 'Active': True}]

        html_template = loader.get_template( load_template )
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:
        html_template = loader.get_template( 'errors/page-404.html' )
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template( 'errors/page-500.html' )
        return HttpResponse(html_template.render(context, request))

###########################################################
# Download Excel Sheet
###########################################################
def object_download(request):
    # Create a workbook and set up the cover sheet
    workbook = Workbook()
    cover_sheet = workbook.active
    cover_sheet.title = "Cover"
    objects_sheet = workbook.create_sheet("Objects")

    # Query all instances of Object_Unit_Download
    elements_list = Object_Unit_Download.objects.all()

    # Populate the Cover sheet
    cover_sheet['A1'] = "XC-PROVIRT TEMPLATE"
    cover_sheet['A2'] = f"Number of elements: {elements_list.count()}"
    cover_sheet['A3'] = f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Populate the Objects sheet with headers
    objects_sheet.append(["ID", "Component", "Version"])
    
    # Iterate over the queryset and write data to the sheet
    for element in elements_list:
        objects_sheet.append([element.id, element.component, element.version])

    # Prepare the HTTP response with the Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="object_download.xlsx"'
    
    # Save the workbook to the response
    workbook.save(response)

    return response